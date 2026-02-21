import json
import yfinance as yf
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# ---------- Setup & Configuration ----------
def load_config():
    with open("config.json") as f:
        return json.load(f)

def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

def apply_heatmap(ws):
    rule = ColorScaleRule(
        start_type="num", start_value=-15, start_color="F8696B", 
        mid_type="num", mid_value=0, mid_color="FFFFFF",         
        end_type="num", end_value=15, end_color="63BE7B"         
    )
    ws.conditional_formatting.add("B2:Z100", rule)

# ---------- Main Processing Loop ----------
def main():
    config = load_config()
    sectors = config["sectors"]
    output_folder = config["output"]["folder"]
    from_date = config.get("from_date", "2023-01-01")

    os.makedirs(output_folder, exist_ok=True)

    for sector, stocks in sectors.items():
        try:
            print(f"🔄 Processing Sector: {sector}...")
            filename = os.path.join(output_folder, f"{sector}.xlsx")

            # 1. Download Data
            raw_data = yf.download(stocks, start=from_date, auto_adjust=True)
            
            if isinstance(raw_data.columns, pd.MultiIndex):
                data = raw_data['Close']
            else:
                data = raw_data['Close'].to_frame(name=stocks[0])
            
            data = data.ffill()

            # 2. Excel Generation
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                data.to_excel(writer, sheet_name="prices")

                # Monthly Returns
                monthly = (data.resample("ME").last().pct_change() * 100).sort_index(ascending=False)
                monthly.index = monthly.index.strftime("%Y-%b")
                monthly.to_excel(writer, sheet_name="monthly_returns")

                # Quarterly Returns
                quarterly = (data.resample("QE").last().pct_change() * 100).sort_index(ascending=False)
                quarterly.index = quarterly.index.to_period("Q").astype(str)
                quarterly.to_excel(writer, sheet_name="quarterly_returns")

                # Stock Rolling Returns (1-Year Window)
                # This requires 'from_date' to be 1 year before your analysis start
                rolling_1y = data.pct_change(periods=252).dropna() * 100
                rolling_1y.to_excel(writer, sheet_name="rolling_12m")

            # 3. Apply Excel Styling
            wb = load_workbook(filename)
            for sheet_name in ["monthly_returns", "quarterly_returns"]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    auto_width(ws)
                    apply_heatmap(ws)
            
            for sheet_name in ["prices", "rolling_12m"]:
                if sheet_name in wb.sheetnames:
                    auto_width(wb[sheet_name])
            
            wb.save(filename)
            print(f"✅ Saved: {filename}")

        except Exception as e:
            print(f"🚨 Error processing {sector}: {str(e)}")

if __name__ == "__main__":
    main()
